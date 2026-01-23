from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.conf import settings
from django.http import JsonResponse, FileResponse, Http404
from django.views.decorators.http import require_POST
from pathlib import Path
import os
import re
import openpyxl
import time
import json
from .excel_manager import get_working_excel_path, sync_new_rows_from_original

# In-memory storage for online users and their current row
# Format: {date: {username: {'row': row_index, 'last_seen': timestamp}}}
online_users = {}
ONLINE_TIMEOUT = 10  # seconds


@login_required
def date_list(request):
    data_root = Path(settings.DATA_ROOT)
    data_file = settings.DATA_FILE
    dates = []

    if data_root.exists():
        for item in sorted(data_root.iterdir(), reverse=True):
            if item.is_dir() and re.match(r'^\d{8}$', item.name):
                has_data = (item / data_file).exists()
                dates.append({
                    'name': item.name,
                    'has_data': has_data
                })

    return render(request, 'data/date_list.html', {'dates': dates})


def get_row_files(date, attribute, seq_num, fits_new, fits_old):
    """Generate file info for a row"""
    data_root = Path(settings.DATA_ROOT)
    data_file = settings.DATA_FILE
    base_dir = data_root / date / Path(data_file).parent

    # Extract base names (remove _new.fits suffix)
    base_new = fits_new.replace('_new.fits', '') if fits_new else None
    base_old = fits_old.replace('_new.fits', '') if fits_old else None

    # Build file prefix: attribute_seqnum_basename
    prefix = f"{attribute}_{seq_num:04d}_"

    files = {'new_time': [], 'old_time': []}

    for base, time_key in [(base_new, 'new_time'), (base_old, 'old_time')]:
        if not base:
            continue
        file_prefix = prefix + base

        # Check for fits files
        for suffix in ['_lib.fits', '_new.fits']:
            fpath = base_dir / (file_prefix + suffix)
            if fpath.exists():
                files[time_key].append({
                    'name': fpath.name,
                    'type': 'fits',
                    'path': str(fpath)
                })

        # Check for jpg files
        for suffix in ['_SEPlib.jpg', '_SEPnew.jpg']:
            fpath = base_dir / (file_prefix + suffix)
            if fpath.exists():
                files[time_key].append({
                    'name': fpath.name,
                    'type': 'jpg',
                    'subtype': 'lib' if 'lib' in suffix else 'new',
                    'path': str(fpath)
                })

    return files


@login_required
def date_detail(request, date):
    rows = []
    headers = []
    error = None
    excel_filename = None

    try:
        # 获取工作Excel文件（仅复制一次）
        file_path = get_working_excel_path(date)
        excel_filename = file_path.name  # 获取文件名

        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active
        header_map = {}
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = list(row)
                header_map = {h: idx for idx, h in enumerate(headers)}
            else:
                rows.append({
                    'index': i,
                    'data': list(row)
                })
        wb.close()
    except FileNotFoundError as e:
        error = f'File not found: {e}'
    except Exception as e:
        error = f'Error: {str(e)}'

    return render(request, 'data/date_detail.html', {
        'date': date,
        'headers': headers,
        'rows': rows,
        'error': error,
        'excel_filename': excel_filename,
        'auto_sync_interval': settings.AUTO_SYNC_INTERVAL
    })


@login_required
def row_files(request, date, row_index):
    """API to get files for a specific row"""
    try:
        file_path = get_working_excel_path(date)

        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active
        headers = []
        target_row = None

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = list(row)
            elif i == row_index:
                target_row = list(row)
                break
        wb.close()

        if not target_row:
            return JsonResponse({'error': 'Row not found'}, status=404)

        # Get column indices
        h_map = {h: idx for idx, h in enumerate(headers)}
        attribute = target_row[h_map.get('attribute', 1)]
        seq_num = target_row[h_map.get('sequence_number', 2)]
        fits_new = target_row[h_map.get('fits_filename_new', 11)]
        fits_old = target_row[h_map.get('fits_filename_old', 15)]
        time_new = target_row[h_map.get('time_utc_new', 10)]
        time_old = target_row[h_map.get('time_utc_old', 14)]

        # Get coordinates
        ra_deg = target_row[h_map.get('ra_deg_new')] if 'ra_deg_new' in h_map else None
        dec_deg = target_row[h_map.get('dec_deg_new')] if 'dec_deg_new' in h_map else None
        ra_hms = target_row[h_map.get('RA_hms_new')] if 'RA_hms_new' in h_map else None
        dec_dms = target_row[h_map.get('Dec_dms_new')] if 'Dec_dms_new' in h_map else None

        files = get_row_files(date, attribute, int(seq_num), fits_new, fits_old)

        # Determine which is earlier
        if time_old and time_new and str(time_old) < str(time_new):
            result = {'left': files['old_time'], 'right': files['new_time'],
                      'left_time': str(time_old), 'right_time': str(time_new)}
        else:
            result = {'left': files['new_time'], 'right': files['old_time'],
                      'left_time': str(time_new), 'right_time': str(time_old)}

        # Add coordinates
        result['ra_deg'] = float(ra_deg) if ra_deg else None
        result['dec_deg'] = float(dec_deg) if dec_deg else None
        result['ra_hms'] = str(ra_hms) if ra_hms else None
        result['dec_dms'] = str(dec_dms) if dec_dms else None

        return JsonResponse(result)
    except FileNotFoundError as e:
        return JsonResponse({'error': str(e)}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def serve_image(request, date, filename):
    """Serve image file"""
    data_root = Path(settings.DATA_ROOT)
    data_file = settings.DATA_FILE
    file_path = data_root / date / Path(data_file).parent / filename

    if not file_path.exists() or not filename.endswith('.jpg'):
        raise Http404("Image not found")

    return FileResponse(open(file_path, 'rb'), content_type='image/jpeg')


@login_required
def serve_fits(request, date, filename):
    """Serve fits file for download"""
    data_root = Path(settings.DATA_ROOT)
    data_file = settings.DATA_FILE
    file_path = data_root / date / Path(data_file).parent / filename

    if not file_path.exists() or not filename.endswith('.fits'):
        raise Http404("File not found")

    response = FileResponse(open(file_path, 'rb'), content_type='application/octet-stream')
    response['Content-Disposition'] = f'attachment; filename="{file_path.name}"'
    return response


def clean_expired_users(date):
    """Remove users who haven't been seen recently"""
    if date not in online_users:
        return
    current_time = time.time()
    expired = [user for user, data in online_users[date].items()
               if current_time - data['last_seen'] > ONLINE_TIMEOUT]
    for user in expired:
        del online_users[date][user]
    if not online_users[date]:
        del online_users[date]


@login_required
@require_POST
def update_status(request, date):
    """Update user's current row status"""
    try:
        data = json.loads(request.body)
        row_index = data.get('row_index')
        username = request.user.username

        if date not in online_users:
            online_users[date] = {}

        online_users[date][username] = {
            'row': row_index,
            'last_seen': time.time()
        }

        clean_expired_users(date)

        return JsonResponse({'status': 'ok'})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def get_status(request, date):
    """Get all online users and their current rows"""
    clean_expired_users(date)

    users = {}
    if date in online_users:
        for username, data in online_users[date].items():
            users[username] = {
                'row': data['row'],
                'last_seen': data['last_seen']
            }

    return JsonResponse({
        'users': users,
        'current_user': request.user.username
    })


# File lock for Excel operations
import threading
excel_lock = threading.Lock()


def get_or_create_judge_column(ws, headers, username):
    """Get or create a judgment column for the user"""
    col_name = f'judge_{username}'
    if col_name in headers:
        return headers.index(col_name) + 1, headers

    # Create new column
    new_col_idx = len(headers) + 1
    ws.cell(row=1, column=new_col_idx, value=col_name)
    headers.append(col_name)
    return new_col_idx, headers


def get_or_create_final_judge_column(ws, headers):
    """Get or create the final judgment column"""
    col_name = 'final_judge'
    if col_name in headers:
        return headers.index(col_name) + 1, headers

    # Create new column at the end
    new_col_idx = len(headers) + 1
    ws.cell(row=1, column=new_col_idx, value=col_name)
    headers.append(col_name)
    return new_col_idx, headers


def get_or_create_final_judge_by_column(ws, headers):
    """Get or create the final judgment by column (records who made the last judgment)"""
    col_name = 'final_judge_by'
    if col_name in headers:
        return headers.index(col_name) + 1, headers

    # Create new column at the end
    new_col_idx = len(headers) + 1
    ws.cell(row=1, column=new_col_idx, value=col_name)
    headers.append(col_name)
    return new_col_idx, headers


def get_or_create_remark_column(ws, headers):
    """Get or create the remark column"""
    col_name = 'final_remark'
    if col_name in headers:
        return headers.index(col_name) + 1, headers

    # Create new column at the end
    new_col_idx = len(headers) + 1
    ws.cell(row=1, column=new_col_idx, value=col_name)
    headers.append(col_name)
    return new_col_idx, headers


@login_required
@require_POST
def submit_judgment(request, date, row_index):
    """Submit a judgment for a row"""
    try:
        file_path = get_working_excel_path(date)
    except FileNotFoundError as e:
        return JsonResponse({'error': str(e)}, status=404)

    try:
        data = json.loads(request.body)
        judgment = data.get('judgment')  # 'exclude', 'suspect', or 'cancel'
        username = request.user.username

        if judgment not in ['exclude', 'suspect', 'cancel']:
            return JsonResponse({'error': 'Invalid judgment'}, status=400)

        with excel_lock:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            # Get headers
            headers = [cell.value for cell in ws[1]]

            # Get or create user's judgment column
            user_col, headers = get_or_create_judge_column(ws, headers, username)

            # Get or create final judgment column
            final_col, headers = get_or_create_final_judge_column(ws, headers)

            # Get or create final judgment by column
            final_by_col, headers = get_or_create_final_judge_by_column(ws, headers)

            if judgment == 'cancel':
                # Clear user's judgment (use empty string instead of None for openpyxl)
                ws.cell(row=row_index + 1, column=user_col).value = ''
                # Clear final judgment
                ws.cell(row=row_index + 1, column=final_col).value = ''
                # Clear who made the final judgment
                ws.cell(row=row_index + 1, column=final_by_col).value = ''
            else:
                # Write user's judgment
                ws.cell(row=row_index + 1, column=user_col, value=judgment)
                # Write final judgment
                ws.cell(row=row_index + 1, column=final_col, value=judgment)
                # Write who made the final judgment
                ws.cell(row=row_index + 1, column=final_by_col, value=username)

            wb.save(file_path)
            wb.close()

        return JsonResponse({
            'status': 'ok',
            'judgment': judgment,
            'user': username
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def get_judgments(request, date):
    """Get all judgments for a date"""
    try:
        file_path = get_working_excel_path(date)
    except FileNotFoundError as e:
        return JsonResponse({'error': str(e)}, status=404)

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]

        # Find judgment and remark columns
        judge_cols = {}
        final_col = None
        final_by_col = None
        remark_col = None
        for idx, h in enumerate(headers):
            if h and h.startswith('judge_'):
                username = h[6:]  # Remove 'judge_' prefix
                judge_cols[username] = idx
            elif h == 'final_judge':
                final_col = idx
            elif h == 'final_judge_by':
                final_by_col = idx
            elif h == 'final_remark':
                remark_col = idx

        # Collect judgments
        judgments = {}
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            row_idx = i + 1
            row_judgments = {}

            for username, col_idx in judge_cols.items():
                if col_idx < len(row) and row[col_idx]:
                    row_judgments[username] = row[col_idx]

            final = None
            if final_col is not None and final_col < len(row):
                final = row[final_col]

            final_by = None
            if final_by_col is not None and final_by_col < len(row):
                final_by = row[final_by_col]

            remark = None
            if remark_col is not None and remark_col < len(row):
                remark = row[remark_col]

            if row_judgments or final or remark:
                judgments[row_idx] = {
                    'users': row_judgments,
                    'final': final,
                    'final_by': final_by,
                    'remark': remark
                }

        wb.close()

        return JsonResponse({
            'judgments': judgments,
            'current_user': request.user.username
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
def submit_remark(request, date, row_index):
    """Submit a remark for a row"""
    try:
        file_path = get_working_excel_path(date)
    except FileNotFoundError as e:
        return JsonResponse({'error': str(e)}, status=404)

    try:
        data = json.loads(request.body)
        remark = data.get('remark', '')

        with excel_lock:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            # Get headers
            headers = [cell.value for cell in ws[1]]

            # Get or create remark column
            remark_col, headers = get_or_create_remark_column(ws, headers)

            # Write remark
            ws.cell(row=row_index + 1, column=remark_col, value=remark)

            wb.save(file_path)
            wb.close()

        return JsonResponse({
            'status': 'ok',
            'remark': remark
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_POST
def sync_excel_rows(request, date):
    """同步原始Excel的新行到工作副本"""
    result = sync_new_rows_from_original(date)

    if result['success']:
        return JsonResponse(result)
    else:
        return JsonResponse(result, status=400)
