# data/excel_manager.py - Excel 文件管理模块

import shutil
from pathlib import Path
from datetime import datetime
from django.conf import settings


def get_working_excel_path(date):
    """
    获取工作用Excel文件路径，仅复制一次

    Args:
        date: 日期字符串 (YYYYMMDD 格式)

    Returns:
        Path: 工作Excel文件的完整路径 (candidate-final-YYYYMMDD-YYYYMMDD-HHmmss.xlsx)

    Raises:
        FileNotFoundError: 原始Excel文件不存在
    """
    data_root = Path(settings.DATA_ROOT)
    data_file = settings.DATA_FILE

    # 原始文件路径
    original_path = data_root / date / data_file

    if not original_path.exists():
        raise FileNotFoundError(f"原始Excel文件不存在: {original_path}")

    # 工作文件目录 (与原始文件同目录)
    excel_dir = original_path.parent

    # 查找是否已存在该日期的工作文件
    existing_file = _find_existing_working_file(excel_dir, date)

    if existing_file:
        return existing_file

    # 生成新的工作文件名（带时间戳）
    now = datetime.now().strftime('%Y%m%d-%H%M%S')
    working_filename = f"candidate-final-{date}-{now}.xlsx"
    working_path = excel_dir / working_filename

    # 复制原始文件一次
    shutil.copy2(original_path, working_path)

    return working_path


def _find_existing_working_file(excel_dir, date):
    """
    查找该日期是否已存在工作文件

    Args:
        excel_dir: Excel文件所在目录
        date: 日期字符串 (YYYYMMDD)

    Returns:
        Path: 找到的工作文件路径，未找到返回None
    """
    pattern = f"candidate-final-{date}-*.xlsx"
    matching_files = list(excel_dir.glob(pattern))

    if matching_files:
        # 返回找到的文件（���该只有一个）
        return matching_files[0]

    return None


def sync_new_rows_from_original(date):
    """
    从原始Excel同步新行到工作副本

    Args:
        date: 日期字符串 (YYYYMMDD 格式)

    Returns:
        dict: {
            'success': bool,
            'added_rows': int,  # 新增的行数
            'total_rows': int,  # 工作副本的总行数（不含表头）
            'message': str
        }
    """
    import openpyxl

    data_root = Path(settings.DATA_ROOT)
    data_file = settings.DATA_FILE

    # 原始文件路径
    original_path = data_root / date / data_file

    if not original_path.exists():
        return {
            'success': False,
            'added_rows': 0,
            'total_rows': 0,
            'message': f'原始Excel文件不存在: {original_path}'
        }

    # 获取工作文件路径
    try:
        working_path = get_working_excel_path(date)
    except FileNotFoundError as e:
        return {
            'success': False,
            'added_rows': 0,
            'total_rows': 0,
            'message': str(e)
        }

    try:
        # 读取原始文件
        wb_original = openpyxl.load_workbook(original_path, read_only=True)
        ws_original = wb_original.active
        original_rows = list(ws_original.iter_rows(values_only=True))
        wb_original.close()

        # 读取工作文件
        wb_working = openpyxl.load_workbook(working_path)
        ws_working = wb_working.active
        working_rows = list(ws_working.iter_rows(values_only=True))

        # 检查表头是否一致（比较数据列）
        if len(original_rows) == 0 or len(working_rows) == 0:
            wb_working.close()
            return {
                'success': False,
                'added_rows': 0,
                'total_rows': len(working_rows) - 1 if working_rows else 0,
                'message': 'Excel文件为空'
            }

        original_header = original_rows[0]
        working_header = working_rows[0]

        # 获取工作副本中的数据列数（原始列数）
        original_col_count = len(original_header)

        # 当前工作副本的数据行数
        current_row_count = len(working_rows) - 1  # 不含表头
        original_row_count = len(original_rows) - 1  # 不含表头

        # 如果原始文件行数 <= 工作副本行数，无需同步
        if original_row_count <= current_row_count:
            wb_working.close()
            return {
                'success': True,
                'added_rows': 0,
                'total_rows': current_row_count,
                'message': f'无新行需要同步（原始: {original_row_count}行, 副本: {current_row_count}行）'
            }

        # 复制新行（从工作副本的下一行开始）
        added_count = 0
        for i in range(current_row_count + 1, original_row_count + 1):
            original_row = original_rows[i]  # 原始文件的第i行（从1开始，0是表头）
            new_row_index = len(working_rows) + 1  # 工作副本的新行号

            # 只复制原始数据列
            for col_idx in range(original_col_count):
                if col_idx < len(original_row):
                    ws_working.cell(row=new_row_index, column=col_idx + 1, value=original_row[col_idx])

            working_rows.append(original_row)  # 更新working_rows列表
            added_count += 1

        # 保存工作文件
        wb_working.save(working_path)
        wb_working.close()

        return {
            'success': True,
            'added_rows': added_count,
            'total_rows': len(working_rows) - 1,
            'message': f'成功同步 {added_count}行新数据'
        }

    except Exception as e:
        return {
            'success': False,
            'added_rows': 0,
            'total_rows': 0,
            'message': f'同步失败: {str(e)}'
        }

